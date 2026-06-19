#!/usr/bin/env python3
"""
A2A Skill Registry Schema v1.0
For Genesis Conductor + Diamondnode QUBO + pbc deployment

This defines the canonical structure for skills in the sovereign AI ecosystem.
"""

from pydantic import BaseModel, Field, field_validator
from typing import List, Optional, Literal, Dict, Any
from datetime import datetime
from enum import Enum


class DeploymentPriority(str, Enum):
    CRITICAL = "Critical"
    HIGH = "High"
    MEDIUM = "Medium"
    LOW = "Low"


class IntegrationLevel(str, Enum):
    NATIVE = "Native"
    HIGH = "High"
    MEDIUM = "Medium"
    LOW = "Low"
    NONE = "None"


class A2ASkill(BaseModel):
    """Canonical A2A Skill entry for Genesis Conductor ecosystem."""

    skill_id: str = Field(..., description="Unique identifier, e.g. FR-POAM-001 or GC-DIAMOND-001")
    skill_name: str = Field(..., description="Human readable name")
    category: str = Field(..., description="e.g. Compliance, Observability, Economics, Orchestration")
    description: str = Field(..., min_length=10)
    version: str = Field(default="1.0.0", description="Semantic version")

    # Economic & Thermodynamic Metrics
    vpd_score: float = Field(..., ge=0, le=100, description="Value Per Diamondnode (0-100)")
    thermodynamic_yield: float = Field(..., ge=0, le=1.0, description="Efficiency vs Landauer limit")
    entropy_production_rate: float = Field(default=0.0, ge=0, description="Normalized EPR")
    dts_score: Optional[float] = Field(default=None, description="Diamondnode Thermodynamic Score (computed)")

    # Deployment & Integration
    diamondnode_ready: bool = Field(default=False)
    deployment_priority: DeploymentPriority = Field(default=DeploymentPriority.MEDIUM)
    genesis_conductor_integration: IntegrationLevel = Field(default=IntegrationLevel.MEDIUM)
    pbc_compatible: bool = Field(default=True)
    a2a_protocol_version: str = Field(default="v1.0")

    # Technical
    dependencies: List[str] = Field(default_factory=list)
    evt_schema_version: Optional[str] = Field(default="1.0")
    mcp_endpoints: List[str] = Field(default_factory=list)
    github_repo: Optional[str] = Field(default=None)
    cloudflare_worker: Optional[str] = Field(default=None)

    # Metadata
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)
    tags: List[str] = Field(default_factory=list)
    owner: str = Field(default="Kovach Enterprises / Genesis Conductor")

    @field_validator('dts_score', mode='before')
    @classmethod
    def compute_dts(cls, v, info):
        """Auto-compute DTS if not provided."""
        if v is not None:
            return v
        data = info.data
        vpd = data.get('vpd_score', 0)
        ty = data.get('thermodynamic_yield', 1.0)
        epr = data.get('entropy_production_rate', 0.0)
        return round(vpd * ty * (1 - min(epr, 1.0)), 4)

    model_config = {
        "json_schema_extra": {
            "example": {
                "skill_id": "FR-POAM-001",
                "skill_name": "FedRAMP POA&M Builder (Next.js)",
                "category": "Compliance",
                "description": "Interactive FedRAMP POA&M generator with Google Drive OAuth and real-time DTS telemetry",
                "vpd_score": 87.0,
                "thermodynamic_yield": 0.92,
                "entropy_production_rate": 0.15,
                "diamondnode_ready": True,
                "deployment_priority": "High",
                "genesis_conductor_integration": "High",
                "pbc_compatible": True,
                "tags": ["fedramp", "compliance", "nextjs", "dts-telemetry"]
            }
        }
    }


class A2ASkillRegistry(BaseModel):
    """Container for multiple A2A Skills with metadata."""
    registry_version: str = "1.0.0"
    generated_at: datetime = Field(default_factory=datetime.utcnow)
    owner: str = "Igor Holt / Kovach Enterprises / Genesis Conductor"
    skills: List[A2ASkill] = Field(default_factory=list)

    def add_skill(self, skill: A2ASkill):
        self.skills.append(skill)
        self.generated_at = datetime.utcnow()

    def to_xlsx(self, filepath: str):
        """Export registry to optimized A2A xlsx format."""
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        data = []
        for skill in self.skills:
            data.append({
                "Skill_ID": skill.skill_id,
                "Skill_Name": skill.skill_name,
                "Category": skill.category,
                "VPD": skill.vpd_score,
                "Thermodynamic_Yield": skill.thermodynamic_yield,
                "DTS": skill.dts_score,
                "Diamondnode_Ready": "Yes" if skill.diamondnode_ready else "No",
                "Deployment_Priority": skill.deployment_priority.value,
                "Genesis_Conductor_Integration": skill.genesis_conductor_integration.value,
                "pbc_Compatible": "Yes" if skill.pbc_compatible else "No",
                "Tags": ", ".join(skill.tags),
            })

        df = pd.DataFrame(data)

        wb = Workbook()
        ws = wb.active
        ws.title = "A2A_Skills_Catalog"

        # Header styling
        header_fill = PatternFill(start_color="0A2540", end_color="0A2540", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for r_idx, row in enumerate(df.values, start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Write headers
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Auto column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 35)

        wb.save(filepath)
        print(f"A2A Skill Registry exported to: {filepath}")


# Example usage
if __name__ == "__main__":
    registry = A2ASkillRegistry(owner="Igor Holt / Genesis Conductor")

    # Add FedRAMP POA&M Builder
    registry.add_skill(A2ASkill(
        skill_id="FR-POAM-001",
        skill_name="FedRAMP POA&M Builder (Next.js)",
        category="Compliance",
        description="Interactive FedRAMP POA&M generator with Google Drive OAuth and real-time DTS telemetry",
        vpd_score=87.0,
        thermodynamic_yield=0.92,
        entropy_production_rate=0.15,
        diamondnode_ready=True,
        deployment_priority=DeploymentPriority.HIGH,
        genesis_conductor_integration=IntegrationLevel.HIGH,
        pbc_compatible=True,
        tags=["fedramp", "compliance", "nextjs", "dts-telemetry", "oauth"]
    ))

    # Add DTS Telemetry
    registry.add_skill(A2ASkill(
        skill_id="GC-DTS-001",
        skill_name="Diamondnode Thermodynamic Score Telemetry",
        category="Observability",
        description="Real-time Landauer-based thermodynamic efficiency telemetry with evt- emission",
        vpd_score=94.0,
        thermodynamic_yield=0.95,
        entropy_production_rate=0.08,
        diamondnode_ready=True,
        deployment_priority=DeploymentPriority.CRITICAL,
        genesis_conductor_integration=IntegrationLevel.NATIVE,
        pbc_compatible=True,
        tags=["diamondnode", "thermodynamics", "telemetry", "evt", "qbo"]
    ))

    registry.to_xlsx("/home/workdir/artifacts/A2A_Skill_Registry_v1.xlsx")
    print("\nJSON Schema:")
    print(A2ASkill.model_json_schema())